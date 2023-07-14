import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.fonts import Font

# 각 항목의 값 별로 매핑되는 정수 값으로 엑셀에 저장
edu = {"기초": ["기초", "기본", "신임", "양성"], "전문": ["전문", "심화"]}
invest = {
    "현장감식": 1,
    "화재감식": 2,
    "범죄분석": 3,
    "거짓말탐지": 4,
    "수중분석": 5,
    "법최면": 6,
    "영상감식": 7,
    "혈흔분석": 8,
    "기타": 9,
}
degree = {"학사": 1, "석사": 2, "박사": 3}

# 전처리 하려는 파일 불러오기
df = pd.read_excel(
    "/Users/nykoh/Desktop/시도청 경력 인적자원정보_통합.xlsx", sheet_name="인적자원정보", header=[0, 1]
)

# 엑셀 행 별로 file_row 객체를 저장하는 변수
rows = []

# 행 데이터 하나 단위의 클래스 구성
class file_row:
    # info = None
    # job_opening = 0
    # basic_edu = {}
    # normal_edu = {}
    # intense_edu = {}
    # invest = {}
    # master = {}
    # degree = {}
    # certificate = {}
    # research = {}
    # patent = {}
    # oversea = {}
    # date_error = False

    def __init__(self, info):
        self.info = info
        self.job_opening = 0
        self.basic_edu = {}
        self.normal_edu = {}
        self.intense_edu = {}
        self.invest = {}
        self.master = {}
        self.degree = {}
        self.certificate = {}
        self.research = {}
        self.patent = {}
        self.oversea = {}
        self.date_error = False

    def print_row(self):
        print(
            "initial:\n",
            "info",
            self.info,
            "\nedu",
            self.basic_edu,
            self.normal_edu,
            self.intense_edu,
            "\ncertficate",
            self.certificate,
            "\ninvers",
            self.invest,
            self.master,
            "\ndegree",
            self.degree,
            "\nresearch",
            self.research,
            self.patent,
            self.oversea,
        )

    # 딕셔너리 업데이트 모듈화
    def update_dict(self, dict, key):
        if key in dict.keys():
            dict[key] = dict[key] + 1
        else:
            dict[key] = 1
        return dict

    # 딕셔너리 별로 업데이트 조건
    def update_edu(self, education, year):
        for e in edu["전문"]:
            if e in education:
                self.intense_edu = self.update_dict(self.intense_edu, year)
                return
        for e in edu["기초"]:
            if e in education:
                self.basic_edu = self.update_dict(self.basic_edu, year)
                return
        self.normal_edu = self.update_dict(self.normal_edu, year)

    def update_invest(self, investment, name, year):
        if investment == "전문수사관":
            try:
                self.invest[year] = invest[name]
            except:
                self.invest[year] = 9  # 기타에 해당하는 경우
        elif investment == "마스터":
            try:
                self.master[year] = invest[name]
            except:
                self.master[year] = 9

    def update_certificate(self, year):
        self.certificate = self.update_dict(self.certificate, year)

    def update_degree(self, degree, year):
        try:
            year = int(year[0:4])
            if degree == "학사":
                self.degree[year] = 1
            elif degree == "석사":
                self.degree[year] = 2
            elif degree == "박사":
                self.degree[year] = 3
            else:
                self.degree[year] = 0
        except:
            pass  # (0by/te/) -> 예외처리

    def update_research(self, year):
        self.research = self.update_dict(self.research, year)

    def update_patent(self, year):
        self.patent = self.update_dict(self.patent, year)

    def update_oversea(self, year):
        self.oversea = self.update_dict(self.oversea, year)

    def check_date_error(self, date):
        if self.job_opening > 0:
            if self.job_opening - date < 0:
                self.date_error = True

    def update_job_opening(self, start, end):
        if start > end:  # 시작일 종료일 같은 경우는 어떡할까요?
            self.date_error = False
        self.check_date_error(end)
        self.job_opening = start

    # 연차별로 항목 개수, 값 저장
    def update_year(self):
        self.job_opening = int(str(self.job_opening)[0:4])
        self.basic_edu = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.basic_edu.items()
        )
        self.normal_edu = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.normal_edu.items()
        )
        self.intense_edu = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.intense_edu.items()
        )
        self.invest = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.invest.items()
        )
        self.master = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.master.items()
        )
        self.degree = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.degree.items()
        )
        self.certificate = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.certificate.items()
        )
        self.patent = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.patent.items()
        )
        self.oversea = dict(
            (max(0, key - self.job_opening + 1), value)
            for (key, value) in self.oversea.items()
        )

    def return_data(self, year):
        data = []
        try:
            if self.basic_edu[year]:
                data.append(self.basic_edu[year])
        except:
            data.append(0)
        try:
            if self.normal_edu[year]:
                data.append(self.normal_edu[year])
        except:
            data.append(0)
        try:
            if self.intense_edu[year]:
                data.append(self.intense_edu[year])
        except:
            data.append(0)
        try:
            if self.invest[year]:
                data.append(self.invest[year])
        except:
            data.append(0)
        try:
            if self.master[year]:
                data.append(self.master[year])
        except:
            data.append(0)
        try:
            if self.certificate[year]:
                data.append(self.certificate[year])
        except:
            data.append(0)
        try:
            if self.degree[year]:
                data.append(self.degree[year])
        except:
            data.append(0)
        try:
            if self.research[year]:
                data.append(self.research[year])
        except:
            data.append(0)
        try:
            if self.patent[year]:
                data.append(self.patent[year])
        except:
            data.append(0)
        try:
            if self.oversea[year]:
                data.append(self.oversea[year])
        except:
            data.append(0)
        return data


new_row = None
for idx, row in df.iterrows():
    # 새로운 사람에 대한 정보가 시작되는 행인 경우
    if type(row[0]) == str:
        if idx > 0:
            # 객체 삽입 후 새로운 객체 생성
            new_row.update_year()
            rows.append(new_row)
        new_row = file_row(row["개인정보"])
        # new_row.print_row()

    if type(row["교육이력"]["기관"]) == str:
        new_row.update_edu(row["교육이력"]["과정명"], int(row["교육이력"]["기간(시작)"][0:4]))

    if type(row["전문수사관"]["수준"]) == str:
        new_row.update_invest(
            row["전문수사관"]["수준"], row["전문수사관"]["분야"], int(row["전문수사관"]["취득일"][0:4])
        )
    if type(row["자격증"]["취득일"]) == str:
        new_row.update_certificate(int(row["자격증"]["취득일"][0:4]))

    if type(row["학력사항"]["구분"]) == str:
        new_row.update_degree(row["학력사항"]["구분"], row["학력사항"]["취득일"])

    if row["공모전/연구실적"]["구분"] == "연구실적":
        new_row.update_research(int(row["공모전/연구실적"]["등록(발표)시기"][0:4]))

    if row["공모전/연구실적"]["구분"] == "특허":
        new_row.update_patent(int(row["공모전/연구실적"]["등록(발표)시기"][0:4]))

    if type(row["해외경력"]["구분"]) == str:
        new_row.update_oversea(int(row["해외경력"]["기간"][0:4]))

    if type(row["과수근무경력"]["기간(시작)"]) == str:
        start_date = int(row["과수근무경력"]["기간(시작)"].replace("/", ""))
        end_date = int(row["과수근무경력"]["기간(종료)"].replace("/", ""))
        new_row.update_job_opening(start_date, end_date)

new_row.update_year()
rows.append(new_row)

# 엑셀 파일 생성
wb = Workbook()
sheet = wb.active

# 칸 넓이 지정
sheet.column_dimensions["A"].width = 57
sheet.column_dimensions["E"].width = 12
sheet.column_dimensions["F"].width = 20

# 개인 정보 이후에 처리되는 정보들을 7열부터 저장
start_column = 7

sheet["A1"] = "개인정보"
sheet.merge_cells("A1:F1")


info_header = ["근무부서", "성별", "나이", "계급", "과수근무총경력", "전문분야1순위"]
data_header = ["교육이력", "", "", "전문수사관", "마스터", "자격증", "학력", "연구실적", "특허", "해외경력"]
edu_header = ["기초", "기본", "심화"]

# 근무 기간 오류 시 해당 데이터 하이라이팅
highlight_fill = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)

for i in range(len(info_header)):
    sheet.cell(row=2, column=i + 1).value = info_header[i]
    sheet.merge_cells(start_row=2, start_column=i + 1, end_row=3, end_column=i + 1)


# 입직 전 ~ 30년차의 기록, 자격증 등의 데이터 저장
for set_idx in range(31):
    sc = start_column + set_idx * 10
    ec = sc + 9
    if set_idx == 0:
        sheet[f"{get_column_letter(sc)}1"] = "입직 전"
    else:
        sheet[f"{get_column_letter(sc)}1"] = f"{set_idx}년차"
    merge_range = f"{get_column_letter(sc)}1:{get_column_letter(ec)}1"
    sheet.merge_cells(merge_range)

    for j in range(len(data_header)):
        sheet.cell(row=2, column=sc + j).value = data_header[j]
    for i in range(len(edu_header)):
        sheet.cell(row=3, column=sc + i).value = edu_header[i]

    sheet.merge_cells(f"{get_column_letter(sc)}2:{get_column_letter(sc+2)}2")
    for i in range(sc + 3, ec + 1):
        sheet.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)

for row_num, row_data in enumerate(rows, start=4):
    for i, d in enumerate(row_data.info, start=1):
        sheet.cell(row=row_num, column=i).value = d
    for i in range(31):
        data = row_data.return_data(i)
        # print("data:", data)
        for j in range(10):
            sheet.cell(row=row_num, column=10 * i + start_column + j).value = data[j]
            sheet.cell(
                row=row_num, column=10 * i + start_column + j
            ).alignment = Alignment(horizontal="center", vertical="center")
    if row_data.date_error:
        for cell in sheet[row_num]:
            cell.fill = highlight_fill

for row in sheet.iter_rows(max_row=3):
    for cell in row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save("police_data.xlsx")
